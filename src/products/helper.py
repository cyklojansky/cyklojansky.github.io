#!../../venv/bin/python
import pandas as pd
import os
import math
import pathlib
import shutil
import argparse
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill, Border, Side
from datetime import datetime

to_download: list = []  # [ [ URL, file_name ] ]

data: dict[list] = {}
dataFrames: dict = {}

group_cols = {}  # number of columns in each group

# categories
col_widths = {"A": 20, "B": 15, "C": 75, "D": 20, "E": 25, "F": 20}
row_heights = 20

update_time = datetime.now().strftime("%d.%m.%Y")


def get_csv_files():
    files = os.listdir()
    return [f for f in files if ".csv" in f]

source_df = pd.read_csv(get_csv_files()[0])


def has_image(sku: str = "", key: str = "", row: int = -1):
    images = os.listdir(f"product-images")
    if sku != "":
        return f"{sku}.webp" in images

    if key != "" and row != -1:
        if key in data.keys():
            return f"{data[key][row]['sku']}.webp" in images
        
    return False
    

def clean():
    os.system("rm seznam-náhradního-spotřebního-materiálu.xlsx")
    os.system("rm seznam-náhradního-spotřebního-materiálu.txt")
    # os.system("rm -rf product-images")


def download_images(folder: str = "product-images"):
    pathlib.Path(folder).mkdir(parents=True, exist_ok=True)
    for url, file_name in to_download:
        extension = url.split(".")[-1]
        full_path = f"{folder}/{file_name}"
        os.system(f"wget \"{url}\" -O \"{full_path}.{extension}\"")
        os.system(f"ffmpeg -i {full_path}.{extension} {full_path}.webp")
        os.system(f"rm {full_path}.{extension}")


def create_picture_link(img: str):
    return f"https://cyklojansky.cz/product-images/{img}.webp"
    #return f"https://cdn.statically.io/gh/cyklojansky/cyklojansky.github.io/gh-pages/product-images/{img}.webp"


def make_excel():
    
    os.system("rm seznam-náhradního-spotřebního-materiálu.xlsx")

    source = source_df.to_dict( orient='records')

    for s in source:
        s: dict[str, str]
        group = ""
        if s["private_note"] == s["private_note"] and s["private_note"].count("_") > 1:
            group = s["private_note"].split("_")[1]

        if group == "":
            continue

        if not group in data.keys():
            data[group] = []
        if s["article_number"] == s["article_number"]:
            s["article_number"] = str(int(s["article_number"]))
        if s["native_retail_price"] == s["native_retail_price"]:
            s["native_retail_price"] = float(s["native_retail_price"])
        s["group"] = group
        data[group].append(s)

    for group, products in data.items():
        d: list = []

        c = 0
        for product in products:
            d.append({"SKU": product["sku"], "Obrázek": create_picture_link(product["sku"]) if has_image(product["sku"]) else "", "Název": f'{product["name"]}', "Cena za jednotku": f'{product["native_retail_price"]} Kč/{product["unit_name"]}', f"Skladem k {update_time}": f'{product["quantity"]} {product["unit_name"]}',
                       "EAN": f'{product["article_number_type"].upper()}: {int(float(product["article_number"]))}' if product["article_number"] == product["article_number"] else ''})
            c += 1

        group_cols[group] = c

        dataFrames[group] = pd.DataFrame(d)

    with pd.ExcelWriter("seznam-náhradního-spotřebního-materiálu.xlsx", engine="openpyxl") as writer:
        workbook = writer.book

        price_style = NamedStyle(name="price_style", number_format='0.0" Kč"', fill=PatternFill(
            start_color="00FF00", end_color="00FF00"))
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )



        # Categories
        for key, df in dataFrames.items():
            df.to_excel(writer, sheet_name=key, index=False)

            sheet = workbook[key]

            # Adjust column widths
            for col, width in col_widths.items():
                sheet.column_dimensions[col].width = width

            for row in range(group_cols[key]+1):
                sheet.row_dimensions[row+1].height = row_heights

            # Add hyperlinks to pictures
            for row in range(group_cols[key]):
                if has_image(key=key, row=row):
                    pic_cell = sheet.cell(row=row+2, column=2)
                    pic_cell.hyperlink = pic_cell.value
                    pic_cell.value = "Odkaz"  # Text to display
                    # Style the hyperlink
                    pic_cell.font = Font(color="000000", underline="single")

            # Alignment
            for y in range(group_cols[key]):
                for x in range(len(col_widths.keys())):
                    sheet[f"{chr(ord('A')+x)}{y+2}"].alignment = Alignment(
                        horizontal="center" if x <= 1 else "left", vertical="center")




        # Calculator


        n = 40
        calc_d = {"SKU": [""]*n,
                  "Název":            [f'=IFERROR(VLOOKUP(TEXT(A{i+2}, "#"),Produkty!A:H, 2, FALSE), "")' for i in range(n)],
                  f"Skladem k {update_time}":          [f'=IFERROR(VLOOKUP(TEXT(A{i+2}, "#"),Produkty!A:H, 7, FALSE), "")' for i in range(n)],
                  "Cena za jednotku": [f'=IFERROR(VLOOKUP(TEXT(A{i+2}, "#"),Produkty!A:H, 4, FALSE), "")' for i in range(n)],
                  "Potřebuji":        [""]*n,
                  "Cena celkem":      [f'=IFERROR(VLOOKUP(TEXT(A{i+2}, \"#\"),Produkty!A:H, 5, FALSE)*E{i+2}, 0)' for i in range(n)], }

        calc_df = pd.DataFrame(calc_d)
        calc_df.to_excel(writer, sheet_name="Kalkulačka nákladů", index=False)

        sheet = workbook["Kalkulačka nákladů"]

        sheet.cell(row=2, column=8).value = "Cena celkem:"
        sheet.cell(row=2, column=9).value = f"=SUM(F2:F{n+1})"

        sheet.cell(row=2, column=9).style = price_style

        for row in range(n):
            sheet.cell(row=row+2, column=6).style = price_style

        # fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # for row in range(n):
        #     sheet.cell(row=row+2, column=1).fill = fill_color

        # Adjust column widths
        for col, width in {"A": 15, "B": 75, "C": 25, "D": 20, "E": 12, "F": 15, "G": 5, "H": 15, "I": 15}.items():
        #for col, width in {"A": 15, "B": 75, "C": 20, "D": 12, "E": 15, "F": 5, "G": 15, "H": 12}.items():
            sheet.column_dimensions[col].width = width

        for row in range(n+1):
            sheet.row_dimensions[row+1].height = row_heights

        for y in range(n):
            if y % 2 == 0:
                fill_color = PatternFill(
                    start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            else:
                fill_color = PatternFill(
                    start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

            if y == 0:
                fill_color = PatternFill(
                    start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

            for x in range(6):
                cell = sheet.cell(row=y+2, column=x+1)
                cell.border = thin_border
                cell.fill = fill_color



        # Products list
        _products = []
        products = []
        for d in list(data.values()):
            _products.extend(d)

        for _p in _products:
            _p: dict

            _ = [_p.pop(key) for key in ['picture'] if key in _p]

            p = {"SKU": _p["sku"],
                 "Group": _p["group"],
                 "Name": _p["name"],
                 "Formated price": f'{_p["native_retail_price"]} Kč/{_p["unit_name"]}',
                 "Price": float(_p["native_retail_price"]),
                 "Unit": _p["unit_name"],
                 "Quantity": f'{_p["quantity"]} {_p["unit_name"]}',
                 "Global number type": _p["article_number_type"],
                 "Global number": _p["article_number"], }

            products.append(p)

        # Products list
        products_df = pd.DataFrame(products)
        products_df.to_excel(writer, sheet_name="Produkty", index=False)
        sheet = workbook["Produkty"]

        # Adjust column widths
        for col, width in {"A": 15, "B": 40, "C": 75, "D": 15, "E": 15, "F": 12, "G": 15, "H": 25, "I": 25}.items():
            sheet.column_dimensions[col].width = width


# make a string of length (more than lenght of the input string)


def lenght(str: str, length: int, mode: str = "e"):
    if mode == "s":
        return ' ' * (length - len(str)) + str
    elif mode == "m":
        l = length - len(str)
        return ' ' * (l/2) + str + ' ' * (l-l/2)
    elif mode == "e":
        return str + (' ' * (length - len(str)))


def make_txt():
    os.system("rm seznam-náhradního-spotřebního-materiálu.txt")
    with open("seznam-náhradního-spotřebního-materiálu.txt", "x+") as f:
        f.write("Cyklo Janský - seznam náhradního spotřebního materiálu\n")
        f.write("https://www.cyklojansky.cz/\n")
        f.write("\n")
        f.write(f"Poslední aktualizace: {update_time}")
        f.write("\n")


        for group, d in data.items():
            f.write("\n")
            f.write(f"==== {group} ====\n")

            for p in d:

                f.write(f"{lenght(p['sku'], 20)}  ") # SKU
                f.write(f"{lenght(p['name'], 70)}  ") # Name
                f.write(f"{lenght(str(float(p['native_retail_price'])).replace(".0", ""), 10, "s")} Kč/{p['unit_name']} ") # Price
                f.write(f"{lenght(str(int(p['quantity'])), 5, "s")} {p['unit_name']} ") # Quantity
                f.write(f"{lenght(p['article_number_type'], 10, "s")}: {p['article_number'] if not math.isnan(float(p['article_number'])) else ''}\n") # EAN


parser = argparse.ArgumentParser(description='Helper for cyklojansky.cz')

parser.add_argument('--excel', action='store_true',
                    help='A boolean flag to make excel file')
parser.add_argument('--txt', action='store_true',
                    help='A boolean flag to make txt file')
parser.add_argument('--download', action='store_true',
                    help='A boolean flag to download images')
parser.add_argument('--clean', action='store_true',
                    help='A boolean flag to clean files')
parser.add_argument('--zip', action='store_true',
                    help='A boolean flag to zip files')
args = parser.parse_args()

if args.clean:
    print("Cleaning files...")
    clean()

if args.excel or args.txt:
    print("Making excel file...")
    make_excel()

if args.txt:
    print("Making txt file...")
    make_txt()


if args.download:
    print("Downloading images...")
    download_images()


if args.zip:
    passwd = input("Password: ")

    os.system(f'zip -P {passwd} seznam-náhradního-spotřebního-materiálu.zip {"seznam-náhradního-spotřebního-materiálu.xlsx" if args.excel else ""} {"seznam-náhradního-spotřebního-materiálu.txt" if args.txt else ""}')

